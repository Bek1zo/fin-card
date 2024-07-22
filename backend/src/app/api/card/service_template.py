import os
import tempfile

import openpyxl
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment

from ...schemas.model import MoneyCertificateRelation
from ...session import session_scope


def select_result(year_one, year_two, status = [False, False]):
    """ Выбор чисел для постановки в итоговое начисление, в зависимости от наличия добавленного второго года """
    if year_two:
        if status[1]:
            return str(str("{:.2f}".format(year_two)).replace(".", ","))
        else:
            return ''
    else:
        if year_one and status[0]:
            return str(str("{:.2f}".format(year_one)).replace(".", ","))
        else:
            return ''


def no_order_name(number):
    """ Возврат ячейки, если отсутствует наименование надбавки """
    result = CellRichText(TextBlock(InlineFont(rFont='Times New Roman', sz=11, u='none', b=True), str(number) + '._______________________________'))
    return result


def filled_order(percentage=None, order_number=None, order_date=None):
    """
    Заполнение ячейки приказа по следующему формату:
    _____ % Пр.№ ________ от "________" ________ 20____ г.
    """
    def _check_order_date(order_date):
        if order_date:
            string = CellRichText(
                TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=True, u='single'), '  ' + str(int(percentage)) + '   ') if percentage is not None else
                TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=False, u='single'), '         '),

                TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=False, u='none'), '%  Пр.№'),
                (TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=True, u='single'), '  ' + order_number + '   ') if order_number is not None else
                 TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=False, u='single'), '          ')),
                TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=False, u='none'), 'от '),
                TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=True, u='single'), order_date.strftime('%d.%m.%Y')),
                TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=False, u='none'), ' г.'))
        else:
            string = CellRichText(

                TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=True, u='single'),
                          '  ' + str(int(percentage)) + '   ') if percentage is not None else
                TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=False, u='single'), '         '),

                TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=False, u='none'), '%  Пр.№'),
                (TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=True, u='single'),
                           '  ' + order_number + '   ') if order_number is not None else
                 TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=False, u='single'), '          ')),
                TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=False, u='none'), 'от '),
                TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=False, u='none'), '"'),
                TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=False, u='single'), '        '),
                TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=False, u='none'), '"'),
                TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=False, u='single'), '         '),
                TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=False, u='none'), '20'),
                TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=False, u='single'), '      '),
                TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=False, u='none'), ' г.'))
        return string
    return _check_order_date(order_date)


# Строка отсутствия приказа
no_order = CellRichText(
    TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=False, u='single'), '         '),
    TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=False, u='none'), '%  Пр.№'),
    TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=False, u='single'), '          '),
    TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=False, u='none'), 'от "'),
    TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=False, u='single'), '        '),
    TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=False, u='none'), '"'),
    TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=False, u='single'), '         '),
    TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=False, u='none'), '20'),
    TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=False, u='single'), '      '),
    TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=False, u='none'), ' г.'))


class blank_payment:
    pnvl_order_number = None
    pnvl_order_date = None
    pnvl_date = None
    pnvl_experience = None
    pnvl_percentage = None
    pnvl_sum = None
    position_salary = None
    rank_salary = None
    ouvs_order_number = None
    ouvs_order_date = None
    ouvs_percentage = None
    ouvs_sum = None
    ouvs_is_active = None
    admission_form_order_number = None
    admission_form_order_date = None
    admission_form_percentage = None
    admission_form_sum = None
    admission_form_is_active = None
    award_order_number = None
    award_order_date = None
    award_percentage = None
    award_sum = None
    award_is_active = None
    qualification_order_number = None
    qualification_order_date = None
    qualification_percentage = None
    qualification_sum = None
    qualification_is_active = None
    primary_premium_order_name = None
    primary_premium_order_number = None
    primary_premium_order_date = None
    primary_premium_percentage = None
    primary_premium_sum = None
    primary_premium_is_active = None
    secondary_premium_order_name = None
    secondary_premium_order_number = None
    secondary_premium_order_date = None
    secondary_premium_percentage = None
    secondary_premium_sum = None
    secondary_premium_is_active = None
    total_salary = None
    total_salary_date = None
    payment_status = None


def get_money_certificate_file(id: str) -> str:
    """ Получение файла карточки ДД """
    with session_scope() as session:
        data = session.query(MoneyCertificateRelation).filter(MoneyCertificateRelation.id == id).order_by(MoneyCertificateRelation.created_utc.desc()).first()

        person = data.person_relation if data.person_relation else {}
        money = data.money_certificate_relation
        date = data.report_period_relation.value.split('-')

        first = data.money_certificate_relation.year_first_relation if data.money_certificate_relation.year_first_relation is not None else blank_payment()
        second = data.money_certificate_relation.year_second_relation if data.money_certificate_relation.year_second_relation is not None else blank_payment()



        """
        Словарь со всеми данными.
        Ключ - Именованная ячейка в шаблоне.
        Значение - Строка или экземпляр класса CellRichText
        """
        serialized_named_range = {}
        ### Шапка ###

        serialized_named_range['year_first'] = date[0]
        serialized_named_range['year_second'] = date[1]
        serialized_named_range['inn'] = person.inn if person.inn else ''
        serialized_named_range['firstname'] = person.firstname if person.firstname else ''
        serialized_named_range['middlename'] = person.middlename if person.middlename else ''
        serialized_named_range['lastname'] = person.lastname if person.lastname else ''
        serialized_named_range['birthday'] = person.birthday.strftime('%d.%m.%Y') if person.birthday else ''
        serialized_named_range['personal_number'] = person.personal_number if person.personal_number else ''
        serialized_named_range['arrived_from'] = person.arrived_from if person.arrived_from else ''

        serialized_named_range['certificate_number'] = money.certificate_number if money.certificate_number else ''
        serialized_named_range['certificate_date'] = money.certificate_date.strftime('%d.%m.%Y') if money.certificate_date else ''
        serialized_named_range['certificate_expire_date'] = money.certificate_expire_date.strftime('%d.%m.%Y') if money.certificate_expire_date else ''

        """
            Данные по ПНВЛ.
            Всегда выбирается второй год, при его наличие.
        """

        if second.pnvl_order_number and second.pnvl_order_date:
            serialized_named_range['pnvl_number_and_order'] = CellRichText(
                TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=False, u='none'), 'Пр. № '),
                TextBlock(InlineFont(rFont='Times New Roman',sz=10, b=True, u='single'), second.pnvl_order_number),
                TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=False, u='none'), ' от '),
                TextBlock(InlineFont(rFont='Times New Roman',sz=10, b=True, u='single'), second.pnvl_order_date.strftime('%d.%m.%Y')),
                TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=False, u='none'), ' г.'))
        elif first.pnvl_order_number and first.pnvl_order_date:
            serialized_named_range['pnvl_number_and_order'] = CellRichText(
                TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=False, u='none'), 'Пр. № '),
                TextBlock(InlineFont(rFont='Times New Roman',sz=10, b=True, u='single'), first.pnvl_order_number),
                TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=False, u='none'), ' от '),
                TextBlock(InlineFont(rFont='Times New Roman',sz=10, b=True, u='single'), first.pnvl_order_date.strftime('%d.%m.%Y')),
                TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=False, u='none'), ' г.'))

        if second.pnvl_date:
            serialized_named_range['pnvl_date'] = CellRichText(
                TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=False, u='none'), 'на '),
                TextBlock(InlineFont(sz=10, b=True, u='single'), second.pnvl_date.strftime('%d.%m.%Y') if second.pnvl_date else ''),
                TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=False, u='none'), ' г.'))
        elif first.pnvl_date:
            serialized_named_range['pnvl_date'] = CellRichText(
                TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=False, u='none'), 'на '),
                TextBlock(InlineFont(sz=10, b=True, u='single'), first.pnvl_date.strftime('%d.%m.%Y') if first.pnvl_date else ''),
                TextBlock(InlineFont(rFont='Times New Rogman', sz=10, b=False, u='none'), ' г.'))

        if second.pnvl_experience:
            serialized_named_range['pnvl_experience'] = CellRichText(
                TextBlock(InlineFont(rFont='Times New Roman',sz=10, b=True, u='single'), '     ' + second.pnvl_experience.split('-')[0] + '     '),
                TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=False, u='none'), 'лет'),
                TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=True, u='single'), '      ' + second.pnvl_experience.split('-')[1] + '     '),
                TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=False, u='none'), 'мес.'),
                TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=True, u='single'), '      ' + second.pnvl_experience.split('-')[2] + '     '),
                TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=False, u='none'), 'дн.'))
        elif first.pnvl_experience:
            serialized_named_range['pnvl_experience'] = CellRichText(
                TextBlock(InlineFont(rFont='Times New Roman',sz=10, b=True, u='single'), '     ' + first.pnvl_experience.split('-')[0] + '     '),
                TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=False, u='none'), 'лет'),
                TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=True, u='single'), '      ' + first.pnvl_experience.split('-')[1] + '     '),
                TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=False, u='none'), 'мес.'),
                TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=True, u='single'), '      ' + first.pnvl_experience.split('-')[2] + '     '),
                TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=False, u='none'), 'дн.'))
        else:
            serialized_named_range['pnvl_experience'] = CellRichText(
                TextBlock(InlineFont(rFont='Times New Roman', sz=10, u='single'), '        '),
                TextBlock(InlineFont(rFont='Times New Roman', sz=10, u='none'), 'лет.'),
                TextBlock(InlineFont(rFont='Times New Roman', sz=10, u='single'), '               '),
                TextBlock(InlineFont(rFont='Times New Roman', sz=10, u='none'), 'мес'),
                TextBlock(InlineFont(rFont='Times New Roman', sz=10, u='single'), '               '),
                TextBlock(InlineFont(rFont='Times New Roman', sz=10, u='none'), 'дн')
            )

        if second.pnvl_percentage:
            serialized_named_range['pnvl_percentage'] = CellRichText(TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=True), str(first.pnvl_percentage) if first.pnvl_percentage else ''))
        elif second.pnvl_percentage:
            serialized_named_range['pnvl_percentage'] = CellRichText(TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=True), str(second.pnvl_percentage) if second.pnvl_percentage else ''))


        """
        Первый год
        Основные надбавки
        """

        if first.ouvs_is_active:
            serialized_named_range['first_ouvs'] = filled_order(first.ouvs_percentage, first.ouvs_order_number, first.ouvs_order_date)
        else:
            serialized_named_range['first_ouvs'] = no_order

        if first.admission_form_is_active:
            serialized_named_range['first_admission'] = filled_order(first.admission_form_percentage, first.admission_form_order_number, first.admission_form_order_date)
        else:
            serialized_named_range['first_admission'] = no_order

        if first.award_is_active:
            serialized_named_range['first_award'] = filled_order(first.award_percentage, first.award_order_number, first.award_order_date)
        else:
            serialized_named_range['first_award'] = no_order

        if first.qualification_is_active:
            serialized_named_range['first_qualification'] = filled_order(first.qualification_percentage, first.qualification_order_number, first.qualification_order_date)
        else:
            serialized_named_range['first_qualification'] = no_order

        """ Дополнительные надбавки """

        if first.primary_premium_is_active:
            serialized_named_range['first_primary_name'] = CellRichText(TextBlock(InlineFont(rFont='Times New Roman', sz=11, u='none', b=True), '5. '),
                                                                        TextBlock(InlineFont(rFont='Times New Roman', sz=11, u='single', b=True), first.primary_premium_order_name))
            serialized_named_range['first_primary'] = filled_order(first.primary_premium_percentage, first.primary_premium_order_number, first.primary_premium_order_date)
        else:
            serialized_named_range['first_primary'] = no_order

        if first.secondary_premium_is_active:
            serialized_named_range['first_secondary_name'] = CellRichText(TextBlock(InlineFont(rFont='Times New Roman', sz=11, u='none', b=True), '6. '), TextBlock(InlineFont(rFont='Times New Roman', sz=11, u='single', b=True), first.secondary_premium_order_name))
            serialized_named_range['first_secondary'] = filled_order(first.secondary_premium_percentage, first.secondary_premium_order_number, first.secondary_premium_order_date)
        else:
            serialized_named_range['first_secondary'] = no_order

        """
        Второй год
        Основные надбавки
        """

        if second.ouvs_is_active:
            serialized_named_range['second_ouvs'] = filled_order(second.ouvs_percentage, second.ouvs_order_number, second.ouvs_order_date)
        else:
            serialized_named_range['second_ouvs'] = no_order

        if second.admission_form_is_active:
            serialized_named_range['second_admission'] = filled_order(second.admission_form_percentage, second.admission_form_order_number, second.admission_form_order_date)
        else:
            serialized_named_range['second_admission'] = no_order

        if second.award_is_active:
            serialized_named_range['second_award'] = filled_order(second.award_percentage, second.award_order_number, second.award_order_date)
        else:
            serialized_named_range['second_award'] = no_order

        if second.qualification_is_active:
            serialized_named_range['second_qualification'] = filled_order(second.qualification_percentage, second.qualification_order_number, second.qualification_order_date)
        else:
            serialized_named_range['second_qualification'] = no_order

        """
        Дополнительные надбавки
        """
        if second.primary_premium_is_active:
            serialized_named_range['second_primary_name'] = CellRichText(
            TextBlock(InlineFont(rFont='Times New Roman', sz=11, u='none', b=True), '5. '),
            TextBlock(InlineFont(rFont='Times New Roman', sz=11, u='single', b=True), second.primary_premium_order_name))
            serialized_named_range['second_primary'] = filled_order(second.primary_premium_percentage, second.primary_premium_order_number, second.primary_premium_order_date)
        else:
            serialized_named_range['second_primary'] = no_order

        if second.secondary_premium_is_active:
            serialized_named_range['second_secondary_name'] = CellRichText(
            TextBlock(InlineFont(rFont='Times New Roman', sz=11, u='none', b=True), '5. '),
            TextBlock(InlineFont(rFont='Times New Roman', sz=11, u='single', b=True), second.secondary_premium_order_name))
            serialized_named_range['second_secondary'] = filled_order(second.secondary_premium_percentage, second.secondary_premium_order_number, second.secondary_premium_order_date)
        else:
            serialized_named_range['second_secondary'] = no_order

        """
        Итоговая выплата, при наличие второго года выбирается он.
        """

        serialized_named_range['pnvl_sum'] = CellRichText(TextBlock(InlineFont(rFont='Times New Roman', sz=11, b=True),
                                           select_result(first.pnvl_sum, second.pnvl_sum, [True, True])))
        serialized_named_range['ouvs_sum'] = CellRichText(TextBlock(InlineFont(rFont='Times New Roman', sz=11, b=True),
                                           select_result(first.ouvs_sum, second.ouvs_sum, [first.ouvs_is_active, second.ouvs_is_active])))
        serialized_named_range['admission_form_sum'] = CellRichText(TextBlock(InlineFont(rFont='Times New Roman', sz=11, b=True),
                                                     select_result(first.admission_form_sum,
                                                                   second.admission_form_sum, [first.admission_form_is_active, second.admission_form_is_active])))
        serialized_named_range['award_sum'] = CellRichText(TextBlock(InlineFont(rFont='Times New Roman', sz=11, b=True),
                                            select_result(first.award_sum, second.award_sum, [first.award_is_active, second.award_is_active])))
        serialized_named_range['qualification_sum'] = CellRichText(TextBlock(InlineFont(rFont='Times New Roman', sz=11, b=True),
                                                    select_result(first.qualification_sum,
                                                                  second.qualification_sum, [first.qualification_is_active, second.qualification_is_active])))
        serialized_named_range['primary_premium_sum'] = CellRichText(TextBlock(InlineFont(rFont='Times New Roman', sz=11, b=True),
                                                      select_result(first.primary_premium_sum,
                                                                    second.primary_premium_sum, [first.primary_premium_is_active, second.primary_premium_is_active])))
        serialized_named_range['secondary_premium_sum'] = CellRichText(TextBlock(InlineFont(rFont='Times New Roman', sz=11, b=True),
                                                        select_result(first.secondary_premium_sum,
                                                                      second.secondary_premium_sum, [first.secondary_premium_is_active, second.secondary_premium_is_active])))
        serialized_named_range['unit'] = CellRichText(
            TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=True), person.unit if person.unit else '')
        )
        serialized_named_range['vus'] = CellRichText(
            TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=True), 'ВУС ' + (person.vus if person.vus else '') + 'Должность ')
        )
        serialized_named_range['job_rank'] = CellRichText(
            TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=True), person.job_rank_relation.name if person.job_rank_relation.name else '')
        )
        serialized_named_range['job_position'] = CellRichText(
            TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=True), person.job_position if person.job_position else '')
        )
        serialized_named_range['tariff_category'] = CellRichText(
            TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=True), person.tariff_category_relation.name if person.tariff_category_relation.name else '')
        )
        serialized_named_range['job_position_order_appointment_number'] = CellRichText(
            TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=False), person.job_position_order_appointment_number if person.job_position_order_appointment_number else '')
        )
        serialized_named_range['job_position_order_appointment_date'] = CellRichText(
            TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=False), str(person.job_position_order_appointment_date.strftime('%d.%m.%Y')) if person.job_position_order_appointment_date else '')
        )
        serialized_named_range['job_position_order_entry_number'] = CellRichText(
            TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=False), person.job_position_order_entry_number if person.job_position_order_entry_number else '')
        )
        serialized_named_range['job_position_order_entry_date'] = CellRichText(
            TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=False), str(person.job_position_order_entry_date.strftime('%d.%m.%Y')) if person.job_position_order_entry_date else '')
        )
        serialized_named_range['position_salary'] = CellRichText(
            TextBlock(InlineFont(rFont='Times New Roman', sz=11, b=True), str(first.position_salary) if first.position_salary else '')
        )
        serialized_named_range['rank_salary'] = CellRichText(
            TextBlock(InlineFont(rFont='Times New Roman', sz=11, b=True), str(first.rank_salary) if first.rank_salary else '')
        )

        serialized_named_range['total_salary'] = CellRichText(
            TextBlock(InlineFont(rFont='Times New Roman', sz=11, b=True), select_result(first.total_salary, second.total_salary, [True, True]))
        )
        serialized_named_range['total_salary_date'] = CellRichText(
            TextBlock(InlineFont(rFont='Times New Roman', sz=10, b=True), str(first.total_salary_date.strftime('%d.%m.%Y')) if first.total_salary_date else '')
        )

        # Текущий каталог, загрузка шаблона, активация первого листа.
        path = os.path.dirname(__file__)

        wb = openpyxl.load_workbook(path + '/template/template.xlsx')

        wb.active = wb['Оборотная']
        ws = wb.active

        ws['A1'] = f'Фамилия, инициалы {person.lastname} {person.firstname[0]}.{person.middlename[0]}.'
        ws['A3'] = date[0]
        ws['J3'] = date[1]

        wb.active = wb['Licevaya']
        ws = wb.active

        # Заполнение данными
        for name, value in serialized_named_range.items():
            if name == 'fullname':
                pass
            cell = wb.defined_names[name]
            cell_position = list(cell.destinations)[0][1].replace('$', '')
            if type(ws[cell_position]) == openpyxl.cell.cell.MergedCell:
                pass
            else:
                cell = ws[cell_position] = value


        # Исправление стилей, после заполнения данных.

        ### Шапка и ПНВЛ ###
        ws['C1'] = '       Личная карточка на денежное довольствие   №_______ за  '
        ws['A13'] = CellRichText(
            TextBlock(InlineFont(rFont='Times New Roman', sz=10, u='single'), '        '),
            TextBlock(InlineFont(rFont='Times New Roman', sz=10, u='none'), 'лет.'),
            TextBlock(InlineFont(rFont='Times New Roman', sz=10, u='single'), '               '),
            TextBlock(InlineFont(rFont='Times New Roman', sz=10, u='none'), 'мес'),
            TextBlock(InlineFont(rFont='Times New Roman', sz=10, u='single'), '               '),
            TextBlock(InlineFont(rFont='Times New Roman', sz=10, u='none'), 'дн')
        )

        #### ПЕРВЫЙ ГОД ####
        # 1. Надбавка за ОУВС
        ws['D11'] = CellRichText(
            TextBlock(InlineFont(rFont='Times New Roman', sz=11, u='none', b=True), '1. '),
            TextBlock(InlineFont(rFont='Times New Roman', sz=11, u='single', b=True), 'Надбавка за ОУВС')) if first.ouvs_is_active else no_order_name(1)
        ws['D13'] = no_order
        ws['O23'] = CellRichText(TextBlock(InlineFont(rFont='Times New Roman', sz=11, u='none', b=True), 'Надбавка за ОУВС')) if first.ouvs_is_active else ''

        # 2. Надбавка за секретность
        ws['D14'] = CellRichText(
            TextBlock(InlineFont(rFont='Times New Roman', sz=11, u='none', b=True), '2. '),
            TextBlock(InlineFont(rFont='Times New Roman', sz=11, u='single', b=True), 'Надбавка за секретность')) if first.admission_form_is_active else no_order_name(2)
        ws['D16'] = no_order
        ws['P23'] = CellRichText(TextBlock(InlineFont(rFont='Times New Roman', sz=11, u='none', b=True), 'Надбавка за секретность')) if first.admission_form_is_active else ''

        # 3. Надбавка за премию
        ws['D17'] = CellRichText(
            TextBlock(InlineFont(rFont='Times New Roman', sz=11, u='none', b=True), '3. '),
            TextBlock(InlineFont(rFont='Times New Roman', sz=11, u='single', b=True), 'Премия')) if first.award_is_active else no_order_name(3)
        ws['D19'] = no_order
        ws['Q23'] = CellRichText(TextBlock(InlineFont(rFont='Times New Roman', sz=11, u='none', b=True), 'Премия')) if first.award_is_active else ''

        # 4. Надбавка за клас. квалификацию
        ws['J11'] = CellRichText(
            TextBlock(InlineFont(rFont='Times New Roman', sz=11, u='none', b=True), '4. '),
            TextBlock(InlineFont(rFont='Times New Roman', sz=11, u='single', b=True), 'Надбавка за клас. квалиф.')) if first.qualification_is_active else no_order_name(4)
        ws['J12'].alignment = Alignment(horizontal='left')
        ws['J13'] = no_order
        ws['J13'].alignment = Alignment(horizontal='left')
        ws['R23'] = CellRichText(TextBlock(InlineFont(rFont='Times New Roman', sz=11, u='none', b=True), 'Надбавка за КК')) if first.qualification_is_active else ''

        # 5. Доп. надбавки первая
        ws['J16'] = no_order

        # 6. Доп. надбавка вторая
        ws['J19'] = no_order

        #### ВТОРОЙ ГОД ####
        # 1. Надбавка за ОУВС
        ws['N11'] = CellRichText(
            TextBlock(InlineFont(rFont='Times New Roman', sz=11, u='none', b=True), '1. '),
            TextBlock(InlineFont(rFont='Times New Roman', sz=11, u='single', b=True), 'Надбавка за ОУВС')) if second.ouvs_percentage else no_order_name(1)
        ws['N13'] = no_order

        # 2. Надбавка за секретность
        ws['N14'] = CellRichText(
            TextBlock(InlineFont(rFont='Times New Roman', sz=11, u='none', b=True), '2. '),
            TextBlock(InlineFont(rFont='Times New Roman', sz=11, u='single', b=True), 'Надбавка за секретность')) if second.admission_form_percentage else no_order_name(2)
        ws['N16'] = no_order

        # 3. Надбавка за премию
        ws['N17'] = CellRichText(
            TextBlock(InlineFont(rFont='Times New Roman', sz=11, u='none', b=True), '3. '),
            TextBlock(InlineFont(rFont='Times New Roman', sz=11, u='single', b=True), 'Премия')) if second.award_percentage else no_order_name(3)
        ws['N19'] = no_order

        # 4. Надбавка за клас. квалификацию
        ws['S11'] = CellRichText(
            TextBlock(InlineFont(rFont='Times New Roman', sz=11, u='none', b=True), '4. '),
            TextBlock(InlineFont(rFont='Times New Roman', sz=11, u='single', b=True), 'Надбавка за клас. квалиф.')) if second.qualification_sum else no_order_name(4)
        ws['S13'] = no_order
        ws['S13'].alignment = Alignment(horizontal='left')

        # 5. Доп. надбавки первая
        ws['S23'] = first.primary_premium_order_name if first.primary_premium_is_active else ''
        ws['S16'] = no_order

        # 6. Доп. надбавка вторая
        ws['S26'] = first.secondary_premium_order_name if first.secondary_premium_is_active else ''
        ws['S19'] = no_order

        # Составитель и проверяющий карточки

        ws['V40'] = money.card_author if money.card_author else ''
        ws['V42'] = money.card_inspector if money.card_inspector else ''

        file_name = f'{tempfile.gettempdir()}/{str(data.person_relation.lastname)}_{str(data.person_relation.firstname)}_{str(data.person_relation.middlename)}_{str(data.report_period_relation.value)}.xlsx'

        wb.save(file_name)

        return file_name
